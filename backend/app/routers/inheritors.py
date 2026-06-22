"""传承人路由"""
import time
from io import BytesIO
from collections import deque
from typing import Optional, List, Tuple, Set

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, status
from sqlalchemy import func, or_
from sqlalchemy.orm import Session
from openpyxl import load_workbook
from pypinyin import lazy_pinyin, Style

from ..database import get_db
from .. import models, schemas
from ..deps import get_current_user
from .genres import _inheritor_to_list

router = APIRouter(prefix="/inheritors", tags=["传承人"])

RELATION_LABELS = {
    "master": "师傅",
    "apprentice": "徒弟",
    "senior_fellow": "师兄/师姐",
    "junior_fellow": "师弟/师妹",
    "sibling": "兄弟姐妹",
    "spouse": "配偶",
    "parent": "父母",
    "child": "子女",
    "cousin": "表/堂兄弟姐妹",
    "uncle_aunt": "叔伯/舅姨",
    "nephew_niece": "侄子/侄女/外甥/外甥女",
    "grandparent": "祖父母/外祖父母",
    "grandchild": "孙辈",
    "colleague": "同事",
    "friend": "友人",
}

RELATION_TYPE_OPTIONS = list(RELATION_LABELS.keys())

MAX_RELATION_DEPTH = 3


def _pinyin_initial(name: str) -> str:
    if not name:
        return ""
    pys = lazy_pinyin(name, style=Style.FIRST_LETTER)
    return pys[0][0].upper() if pys and pys[0] else ""


def _build_query(db: Session, genre: Optional[int] = None, role: Optional[str] = None,
                 region: Optional[str] = None, master: Optional[int] = None,
                 age_group: Optional[str] = None, pinyin: Optional[str] = None,
                 keyword: Optional[str] = None):
    q = db.query(models.Inheritor)
    if genre:
        q = q.filter(models.Inheritor.genre_id == genre)
    if role:
        q = q.filter(models.Inheritor.role_type == role)
    if region:
        q = q.filter(models.Inheritor.region.like(f"%{region}%"))
    if master:
        q = q.filter(models.Inheritor.master_id == master)
    if age_group:
        q = q.filter(models.Inheritor.age_group == age_group)
    if pinyin:
        q = q.filter(models.Inheritor.pinyin_initial == pinyin.upper())
    if keyword:
        # 走 pg_trgm GIN 索引
        like = f"%{keyword}%"
        q = q.filter(or_(
            models.Inheritor.name.ilike(like),
            models.Inheritor.region.ilike(like),
        ))
    return q


@router.get("", response_model=schemas.Page[schemas.InheritorListItem])
def list_inheritors(
    genre: Optional[int] = Query(None),
    role: Optional[str] = Query(None),
    region: Optional[str] = Query(None),
    master: Optional[int] = Query(None),
    age_group: Optional[str] = Query(None),
    pinyin: Optional[str] = Query(None),
    keyword: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    size: int = Query(12, ge=1, le=500),
    db: Session = Depends(get_db),
):
    q = _build_query(db, genre, role, region, master, age_group, pinyin, keyword)
    total = q.count()
    items = (
        q.order_by(func.coalesce(models.Inheritor.pinyin_initial, 'Z').asc(),
                   models.Inheritor.name.asc())
         .offset((page - 1) * size)
         .limit(size)
         .all()
    )
    return {
        "items": [_inheritor_to_list(i) for i in items],
        "total": total,
        "page": page,
        "size": size,
    }


@router.get("/search", response_model=list[schemas.InheritorListItem])
def search_inheritors(keyword: str = Query(..., min_length=1), db: Session = Depends(get_db)):
    """模糊搜索姓名（使用 pg_trgm GIN 索引，<300ms）"""
    like = f"%{keyword}%"
    items = (
        db.query(models.Inheritor)
          .filter(models.Inheritor.name.ilike(like))
          .limit(50)
          .all()
    )
    return [_inheritor_to_list(i) for i in items]


@router.get("/pinyin/{initial}", response_model=list[schemas.InheritorListItem])
def inheritors_by_pinyin(initial: str, db: Session = Depends(get_db)):
    items = (
        db.query(models.Inheritor)
          .filter(models.Inheritor.pinyin_initial == initial.upper())
          .order_by(models.Inheritor.name.asc())
          .all()
    )
    return [_inheritor_to_list(i) for i in items]


@router.get("/{inheritor_id}", response_model=schemas.InheritorDetail)
def get_inheritor(inheritor_id: int, db: Session = Depends(get_db)):
    ih = db.query(models.Inheritor).filter(models.Inheritor.id == inheritor_id).first()
    if not ih:
        raise HTTPException(status_code=404, detail="传承人不存在")
    base = _inheritor_to_list(ih)
    learning = [schemas.Learning(id=l.id, year=l.year, title=l.title, description=l.description) for l in ih.learning]
    plays = [schemas.InheritorPlay(
        id=p.id, play_id=p.play_id, play_name=p.play.name if p.play else None, role_name=p.role_name
    ) for p in ih.plays_link]
    awards = [schemas.Award(id=a.id, name=a.name, year=a.year, level=a.level) for a in ih.awards]
    media = [schemas.Media(id=m.id, type=m.type, title=m.title, file_path=m.file_path, description=m.description) for m in ih.media]
    return schemas.InheritorDetail(
        **base.model_dump(),
        learning=learning,
        plays=plays,
        awards=awards,
        media=media,
    )


@router.post("", response_model=schemas.InheritorDetail, status_code=status.HTTP_201_CREATED)
def create_inheritor(data: schemas.InheritorInput, db: Session = Depends(get_db),
                     _user=Depends(get_current_user)):
    ih = models.Inheritor(
        **data.model_dump(exclude_unset=True),
        pinyin_initial=_pinyin_initial(data.name),
    )
    db.add(ih)
    db.commit()
    db.refresh(ih)
    return get_inheritor(ih.id, db)


@router.put("/{inheritor_id}", response_model=schemas.InheritorDetail)
def update_inheritor(inheritor_id: int, data: schemas.InheritorInput,
                     db: Session = Depends(get_db), _user=Depends(get_current_user)):
    ih = db.query(models.Inheritor).filter(models.Inheritor.id == inheritor_id).first()
    if not ih:
        raise HTTPException(status_code=404, detail="传承人不存在")
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(ih, k, v)
    if data.name:
        ih.pinyin_initial = _pinyin_initial(data.name)
    db.commit()
    db.refresh(ih)
    return get_inheritor(ih.id, db)


@router.delete("/{inheritor_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_inheritor(inheritor_id: int, db: Session = Depends(get_db),
                     _user=Depends(get_current_user)):
    ih = db.query(models.Inheritor).filter(models.Inheritor.id == inheritor_id).first()
    if not ih:
        raise HTTPException(status_code=404, detail="传承人不存在")
    db.delete(ih)
    db.commit()
    return None


@router.post("/batch", response_model=schemas.BatchImportResult)
def batch_import(file: UploadFile = File(...), db: Session = Depends(get_db),
                 _user=Depends(get_current_user)):
    """批量导入 Excel。列：姓名、性别、出生年月、年龄段、剧种、行当、地区、师承、简介"""
    start = time.time()
    content = file.file.read()
    wb = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return schemas.BatchImportResult(total=0, success=0, failed=0, errors=[], elapsed_ms=0)

    # 查找表头行
    header_idx = 0
    headers = [str(c).strip() if c else "" for c in rows[0]]
    # 列名映射（支持中英文与多种写法）
    aliases = {
        "name": ["姓名", "名字", "传承人姓名"],
        "gender": ["性别"],
        "birth": ["出生年月", "出生日期", "生日"],
        "age": ["年龄段", "年龄阶段"],
        "genre": ["剧种", "所属剧种"],
        "role": ["行当", "角色", "工", "工"],
        "region": ["地区", "籍贯", "所在地"],
        "master": ["师承", "师父", "师傅"],
        "bio": ["简介", "个人简介", "介绍"],
    }

    def find_col(names):
        for n in names:
            if n in headers:
                return headers.index(n)
        return -1

    col = {k: find_col(v) for k, v in aliases.items()}
    data_rows = rows[1:]

    # 预取剧种映射
    genre_map = {g.name: g for g in db.query(models.Genre).all()}

    errors: list[str] = []
    new_records: list[models.Inheritor] = []
    processed_names: set[str] = set()

    for idx, row in enumerate(data_rows, start=2):
        if not row or all(v in (None, "") for v in row):
            continue
        try:
            def val(key):
                return row[col[key]] if col[key] >= 0 and col[key] < len(row) else None

            name = str(val("name") or "").strip()
            if not name:
                errors.append(f"第{idx}行：姓名为空，已跳过")
                continue
            if name in processed_names:
                errors.append(f"第{idx}行：姓名'{name}'重复，已跳过")
                continue
            processed_names.add(name)

            genre_name = str(val("genre") or "").strip() if col["genre"] >= 0 else ""
            genre = genre_map.get(genre_name) if genre_name else None
            if genre_name and not genre:
                errors.append(f"第{idx}行：剧种'{genre_name}'不存在，已跳过")
                continue

            master_name = str(val("master") or "").strip() if col["master"] >= 0 else ""
            master_id = None
            if master_name:
                # 同批次或已存在中查找
                existing = (
                    db.query(models.Inheritor)
                      .filter(models.Inheritor.name == master_name)
                      .first()
                )
                if existing:
                    master_id = existing.id

            gender = str(val("gender") or "").strip() if col["gender"] >= 0 else None
            birth_raw = val("birth") if col["birth"] >= 0 else None
            birth_date = None
            if birth_raw:
                if hasattr(birth_raw, "isoformat"):
                    birth_date = birth_raw
                else:
                    try:
                        s = str(birth_raw).strip().replace("/", "-").replace(".", "-")
                        # 截取 YYYY-MM-DD
                        from datetime import date as date_mod
                        parts = s.split("-")
                        if len(parts) >= 3:
                            birth_date = date_mod(int(parts[0]), int(parts[1]), int(parts[2][:2]))
                    except Exception:
                        birth_date = None

            age_group = str(val("age") or "").strip() if col["age"] >= 0 else None
            role_type = str(val("role") or "").strip() if col["role"] >= 0 else None
            region = str(val("region") or "").strip() if col["region"] >= 0 else None
            bio = str(val("bio") or "").strip() if col["bio"] >= 0 else None

            new_records.append(models.Inheritor(
                name=name,
                pinyin_initial=_pinyin_initial(name),
                gender=gender or None,
                birth_date=birth_date,
                age_group=age_group or None,
                genre_id=genre.id if genre else None,
                role_type=role_type or None,
                region=region or None,
                master_id=master_id,
                bio=bio or None,
            ))
        except Exception as e:
            errors.append(f"第{idx}行：{e}")

    # 批量插入
    success = 0
    if new_records:
        db.bulk_save_objects(new_records)
        db.commit()
        success = len(new_records)

    elapsed_ms = int((time.time() - start) * 1000)
    return schemas.BatchImportResult(
        total=len(data_rows),
        success=success,
        failed=len(data_rows) - success,
        errors=errors[:100],
        elapsed_ms=elapsed_ms,
    )


def _get_label(rtype: str) -> str:
    return RELATION_LABELS.get(rtype, rtype)


def _find_shared_learning(db: Session, a_id: int, b_id: int) -> Optional[str]:
    """查找两位传承人的共同学习经历"""
    a_years = {(l.year, l.title) for l in db.query(models.Learning).filter(models.Learning.inheritor_id == a_id).all()}
    b_years = {(l.year, l.title) for l in db.query(models.Learning).filter(models.Learning.inheritor_id == b_id).all()}
    common = a_years & b_years
    if not common:
        return None
    items = [f"{y}年《{t}》" for y, t in sorted(common) if y and t]
    return "、".join(items) if items else None


def _collect_immediate_relations(
    db: Session,
    center_id: int,
    visited: Set[int],
) -> List[Tuple[int, str, str, int]]:
    """收集某个传承人的直接关系，返回 [(target_id, relation_type, direction, distance)]"""
    results: List[Tuple[int, str, str, int]] = []

    center = db.query(models.Inheritor).filter(models.Inheritor.id == center_id).first()
    if not center:
        return results

    # 1. 师徒关系（来自 master_id 字段）
    if center.master_id and center.master_id not in visited:
        results.append((center.master_id, "master", "up", 1))
    # 徒弟（查找 master_id = center.id 的人）
    apprentices = (
        db.query(models.Inheritor)
          .filter(models.Inheritor.master_id == center_id)
          .all()
    )
    for app in apprentices:
        if app.id not in visited:
            results.append((app.id, "apprentice", "down", 1))
    # 同门（同一个师傅的其他人）
    if center.master_id:
        fellows = (
            db.query(models.Inheritor)
              .filter(
                  models.Inheritor.master_id == center.master_id,
                  models.Inheritor.id != center_id,
              )
              .all()
        )
        for f in fellows:
            if f.id not in visited:
                # 粗略按出生年分师兄师弟
                if center.birth_date and f.birth_date and f.birth_date < center.birth_date:
                    results.append((f.id, "senior_fellow", "side", 2))
                else:
                    results.append((f.id, "junior_fellow", "side", 2))

    # 2. 多维度关系（来自 inheritor_relations 表）
    out_rels = (
        db.query(models.InheritorRelation)
          .filter(models.InheritorRelation.from_inheritor_id == center_id)
          .all()
    )
    for r in out_rels:
        if r.to_inheritor_id not in visited:
            results.append((r.to_inheritor_id, r.relation_type, "out", 1))

    in_rels = (
        db.query(models.InheritorRelation)
          .filter(models.InheritorRelation.to_inheritor_id == center_id)
          .all()
    )
    for r in in_rels:
        if r.from_inheritor_id not in visited:
            # 反向映射关系类型
            rev_map = {
                "parent": "child",
                "child": "parent",
                "uncle_aunt": "nephew_niece",
                "nephew_niece": "uncle_aunt",
                "grandparent": "grandchild",
                "grandchild": "grandparent",
                "senior_fellow": "junior_fellow",
                "junior_fellow": "senior_fellow",
            }
            rev_type = rev_map.get(r.relation_type, r.relation_type)
            results.append((r.from_inheritor_id, rev_type, "in", 1))

    return results


@router.get("/{inheritor_id}/relations", response_model=schemas.InheritorRelationsResponse)
def get_inheritor_relations(
    inheritor_id: int,
    relation_type: Optional[str] = Query(None, description="按关系类型筛选"),
    max_distance: int = Query(MAX_RELATION_DEPTH, ge=1, le=5, description="最大关联距离"),
    db: Session = Depends(get_db),
):
    """获取传承人的所有关联人（含师徒、亲属、同门等）"""
    center = db.query(models.Inheritor).filter(models.Inheritor.id == inheritor_id).first()
    if not center:
        raise HTTPException(status_code=404, detail="传承人不存在")

    # BFS 遍历，避免重复
    visited: Set[int] = {inheritor_id}
    queue: deque = deque()
    queue.append((inheritor_id, 0))

    relation_items: List[schemas.InheritorRelationItem] = []

    while queue:
        current_id, current_dist = queue.popleft()
        if current_dist >= max_distance:
            continue

        immediate = _collect_immediate_relations(db, current_id, visited)

        for target_id, rtype, direction, edge_dist in immediate:
            if target_id in visited:
                continue
            visited.add(target_id)

            total_dist = current_dist + edge_dist
            if total_dist > max_distance:
                continue

            # 按关系类型过滤（第一层直接匹配，间接关系不过滤类型但保留）
            if relation_type and current_dist == 0 and rtype != relation_type:
                continue

            target_ih = db.query(models.Inheritor).filter(models.Inheritor.id == target_id).first()
            if not target_ih:
                continue

            shared = _find_shared_learning(db, inheritor_id, target_id) if total_dist <= 2 else None

            relation_items.append(schemas.InheritorRelationItem(
                inheritor=_inheritor_to_list(target_ih),
                relation_type=rtype,
                relation_label=_get_label(rtype),
                distance=total_dist,
                shared_learning=shared,
                direction=direction,
            ))

            if total_dist < max_distance:
                queue.append((target_id, total_dist))

    return schemas.InheritorRelationsResponse(
        center=_inheritor_to_list(center),
        relations=sorted(relation_items, key=lambda x: (x.distance, x.relation_label)),
    )
